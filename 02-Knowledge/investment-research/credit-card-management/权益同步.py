"""权益统计数据同步脚本 — 从 信用卡主控表.xlsx 的 4-权益变现 Sheet 同步到 信用卡管理方案.md

=== 同步方向 ===
  Excel (主数据源) → Markdown (只读展示)

=== 使用方式 ===
  1. 手动: python 权益同步.py
  2. 自动: 加入年费到期提醒流程，每次运行提醒时顺便同步

=== 同步内容 ===
  - 4A. 权益库存表（库存数据 + 合计统计）
  - 4B. 变现流水（流水记录）
"""

import openpyxl
import datetime
import re
from pathlib import Path

# ============================================
# 配置
# ============================================

EXCEL_PATH = Path(__file__).parent / "信用卡主控表.xlsx"
MD_PATH = Path(__file__).parent / "信用卡管理方案.md"

# ============================================
# 读取 Excel 权益数据
# ============================================

def read_benefit_inventory(ws):
    """读取 4A. 权益库存（Row 5-15）"""
    headers = [ws.cell(row=4, column=c).value for c in range(1, 9)]
    inventory = []
    for row in range(5, 16):
        bank = ws.cell(row=row, column=1).value
        if not bank or bank in ("合计",):
            continue
        item = {}
        for c in range(1, 9):
            val = ws.cell(row=row, column=c).value
            key = headers[c - 1] if headers[c - 1] else f"col{c}"
            item[key] = str(val) if val is not None else "—"
        inventory.append(item)

    # 从数据行计算合计（避免读取公式单元格的公式字符串）
    total = {
        "银行": "合计",
        "权益": "",
        "年度总量": sum(int(r["年度总量"]) for r in inventory if r["年度总量"].isdigit()),
        "已卖出": sum(int(r["已卖出"]) for r in inventory if r["已卖出"].isdigit()),
        "剩余": sum(int(r["剩余"]) for r in inventory if r["剩余"].isdigit()),
    }
    # 预估收入合计：从 Excel 合计行读取（F17:G17 合并单元格）
    f17 = ws.cell(row=17, column=6).value
    total["预估收入"] = str(f17) if f17 else "—"

    return headers, inventory, total

def read_benefit_ledger(ws):
    """读取 4B. 变现流水（Row 20+）"""
    headers = [ws.cell(row=20, column=c).value for c in range(1, 8)]
    ledger = []
    for row in range(21, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        if not date_val or date_val == "—":
            continue
        item = {}
        for c in range(1, 8):
            val = ws.cell(row=row, column=c).value
            key = headers[c - 1] if headers[c - 1] else f"col{c}"
            if isinstance(val, datetime.datetime):
                item[key] = val.strftime("%Y-%m-%d")
            else:
                item[key] = str(val) if val is not None else "—"
        ledger.append(item)
    return headers, ledger

# ============================================
# 生成 Markdown 表格
# ============================================

def build_markdown_table(data, headers=None, cols=None):
    """通用 Markdown 表格生成"""
    if not data:
        return ""
    if cols is None:
        cols = list(data[0].keys())
    if headers is None:
        headers = cols

    lines = []
    lines.append("| " + " | ".join(str(h) for h in headers) + " |")
    lines.append("|" + "|".join("------" for _ in headers) + "|")
    for row in data:
        lines.append("| " + " | ".join(str(row.get(c, "—")) for c in cols) + " |")
    return "\n".join(lines)

def generate_inventory_section(headers, inventory, total):
    """生成 4A. 权益库存 Markdown"""
    cols = ["银行", "权益", "年度总量", "已卖出", "剩余", "市场价(元)", "预估收入", "限制"]
    lines = ["### Sheet 4：权益变现", "", "**4A. 权益库存（可卖权益汇总）**", ""]
    lines.append(build_markdown_table(inventory, headers=cols))

    # 合计行
    lines.append("")
    total_text = f"**预估年度收入合计：{total.get('预估收入', '—')}**"
    lines.append(total_text)

    return "\n".join(lines)

def generate_ledger_section(headers, ledger):
    """生成 4B. 变现流水 Markdown"""
    cols = ["日期", "银行", "权益", "售价(元)", "买家", "状态", "备注"]
    lines = ["", "**4B. 变现流水**", ""]
    if ledger:
        lines.append(build_markdown_table(ledger, headers=cols))
    else:
        lines.append("| 日期 | 银行 | 权益 | 售价(元) | 买家 | 状态 | 备注 |")
        lines.append("|------|------|------|----------|------|------|------|")
        lines.append("| — | — | — | — | — | — | — |")
    return "\n".join(lines)

# ============================================
# 更新 Markdown 文件
# ============================================

def update_markdown(new_section, md_path):
    """
    在 md 文件中定位并替换 `### Sheet 4：权益变现` 开始的整个区块。
    替换到文件末尾或下一个 `### ` 标题（`### Sheet 5`）之前。
    """
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()

    # 定位 Sheet 4 区块
    start_marker = "### Sheet 4：权益变现"
    end_marker = "\n### Sheet 5"

    start_pos = content.find(start_marker)
    if start_pos == -1:
        print(f"❌ 未找到 `{start_marker}` 标记，无法同步")
        return False

    end_pos = content.find(end_marker, start_pos)
    if end_pos == -1:
        # 到文件末尾
        end_pos = len(content)

    # 替换
    new_content = content[:start_pos] + new_section + content[end_pos:]

    with open(md_path, "w", encoding="utf-8") as f:
        f.write(new_content)

    print(f"✅ 已同步权益数据到 {md_path.name}")
    return True

# ============================================
# 主入口
# ============================================

def main():
    if not EXCEL_PATH.exists():
        print(f"❌ Excel 文件不存在: {EXCEL_PATH}")
        return
    if not MD_PATH.exists():
        print(f"❌ Markdown 文件不存在: {MD_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["4-权益变现"]

    # 读取数据
    inv_headers, inventory, total = read_benefit_inventory(ws)
    led_headers, ledger = read_benefit_ledger(ws)
    wb.close()

    # 生成内容
    inv_section = generate_inventory_section(inv_headers, inventory, total)
    led_section = generate_ledger_section(led_headers, ledger)
    full_section = inv_section + "\n" + led_section

    # 打印预览
    print(f"=== 权益数据同步 === {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"\n📊 权益库存共 {len(inventory)} 条")
    print(f"💰 预估收入合计: {total.get('预估收入', '—')}")
    if ledger:
        print(f"📝 变现流水共 {len(ledger)} 条")

    # 写入 Markdown
    update_markdown(full_section, MD_PATH)

if __name__ == "__main__":
    main()