"""同步权益与收支数据到信用卡主控表.xlsx"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="微软雅黑", size=10)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(name="微软雅黑", bold=True, size=10)

input_path = r"e:\ProjectGroup\AI\ContextStack\02-Knowledge\investment-research\credit-card-management\信用卡主控表.xlsx"
wb = openpyxl.load_workbook(input_path)

def is_merged(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return True
    return False

# ================================================
# 1. Sheet 4 权益变现 — 补全权益数据
# ================================================
ws4 = wb["4-权益变现"]

# 当前权益行：行5-15（11行），合计行：行16
# 需要在邮储机场贵宾厅（行9）后插入邮储高铁贵宾厅
# 在行9之后插入一行
ws4.insert_rows(10)

# 填写新行数据
new_row = 10
ws4.cell(row=new_row, column=1, value="邮储").font = DATA_FONT
ws4.cell(row=new_row, column=1).alignment = center_align
ws4.cell(row=new_row, column=1).border = thin_border
ws4.cell(row=new_row, column=2, value="高铁贵宾厅").font = DATA_FONT
ws4.cell(row=new_row, column=2).alignment = center_align
ws4.cell(row=new_row, column=2).border = thin_border
ws4.cell(row=new_row, column=3, value=6).font = DATA_FONT
ws4.cell(row=new_row, column=3).alignment = center_align
ws4.cell(row=new_row, column=3).border = thin_border
ws4.cell(row=new_row, column=4, value=0).font = DATA_FONT
ws4.cell(row=new_row, column=4).alignment = center_align
ws4.cell(row=new_row, column=4).border = thin_border
ws4.cell(row=new_row, column=5, value=6).font = DATA_FONT
ws4.cell(row=new_row, column=5).alignment = center_align
ws4.cell(row=new_row, column=5).border = thin_border
ws4.cell(row=new_row, column=6, value="50-80").font = DATA_FONT
ws4.cell(row=new_row, column=6).alignment = center_align
ws4.cell(row=new_row, column=6).border = thin_border
ws4.cell(row=new_row, column=7, value="300-480").font = DATA_FONT
ws4.cell(row=new_row, column=7).alignment = center_align
ws4.cell(row=new_row, column=7).border = thin_border
ws4.cell(row=new_row, column=8, value="可吃饭,可带人").font = DATA_FONT
ws4.cell(row=new_row, column=8).alignment = center_align
ws4.cell(row=new_row, column=8).border = thin_border

# 更新合计行（现在合计行在 17）
total_row = 17
# 合并 A-E
ws4.merge_cells(f"A{total_row}:E{total_row}")
ws4.cell(row=total_row, column=1, value="合计").font = TOTAL_FONT
ws4.merge_cells(f"F{total_row}:G{total_row}")
ws4.cell(row=total_row, column=6, value="3,150-5,468 元").font = TOTAL_FONT
for c in range(1, 9):
    ws4.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws4.cell(row=total_row, column=c).border = thin_border
    ws4.cell(row=total_row, column=c).alignment = center_align

print("✅ Sheet 4 权益变现：已补全邮储高铁贵宾厅（6次/年）")

# ================================================
# 2. Sheet 6 收支总账 — 同步最新年费数据
# ================================================
ws6 = wb["6-收支总账"]

# 更新 6A 支出表，添加更多明细
# 当前支出表行5-7（中信/交通/兴业），合计行8
# 在第8行前插入2行（光大+宁波）
ws6.insert_rows(8, 3)  # 插入3行

# 第8行：光大
ws6.cell(row=8, column=1, value="光大").font = DATA_FONT
ws6.cell(row=8, column=1).alignment = center_align; ws6.cell(row=8, column=1).border = thin_border
ws6.cell(row=8, column=2, value="阳光车主").font = DATA_FONT
ws6.cell(row=8, column=2).alignment = center_align; ws6.cell(row=8, column=2).border = thin_border
ws6.cell(row=8, column=3, value=500).font = DATA_FONT
ws6.cell(row=8, column=3).alignment = center_align; ws6.cell(row=8, column=3).border = thin_border
ws6.cell(row=8, column=4, value="积分抵扣").font = DATA_FONT
ws6.cell(row=8, column=4).alignment = center_align; ws6.cell(row=8, column=4).border = thin_border
ws6.cell(row=8, column=5, value="—").font = DATA_FONT
ws6.cell(row=8, column=5).alignment = center_align; ws6.cell(row=8, column=5).border = thin_border
ws6.cell(row=8, column=6, value="10W积分").font = DATA_FONT
ws6.cell(row=8, column=6).alignment = center_align; ws6.cell(row=8, column=6).border = thin_border

# 第9行：宁波
ws6.cell(row=9, column=1, value="宁波").font = DATA_FONT
ws6.cell(row=9, column=1).alignment = center_align; ws6.cell(row=9, column=1).border = thin_border
ws6.cell(row=9, column=2, value="菁英卡").font = DATA_FONT
ws6.cell(row=9, column=2).alignment = center_align; ws6.cell(row=9, column=2).border = thin_border
ws6.cell(row=9, column=3, value=1800).font = DATA_FONT
ws6.cell(row=9, column=3).alignment = center_align; ws6.cell(row=9, column=3).border = thin_border
ws6.cell(row=9, column=4, value="消费额度").font = DATA_FONT
ws6.cell(row=9, column=4).alignment = center_align; ws6.cell(row=9, column=4).border = thin_border
ws6.cell(row=9, column=5, value="—").font = DATA_FONT
ws6.cell(row=9, column=5).alignment = center_align; ws6.cell(row=9, column=5).border = thin_border
ws6.cell(row=9, column=6, value="18W").font = DATA_FONT
ws6.cell(row=9, column=6).alignment = center_align; ws6.cell(row=9, column=6).border = thin_border

# 第10行：空行分隔
ws6.cell(row=10, column=1, value="").font = DATA_FONT
ws6.cell(row=10, column=1).border = thin_border

# 第11行：合计行（重新计算）
total_row_6a = 11
ws6.merge_cells(f"A{total_row_6a}:B{total_row_6a}")
ws6.cell(row=total_row_6a, column=1, value="合计").font = TOTAL_FONT
ws6.cell(row=total_row_6a, column=3, value="=SUM(C5:C9)").font = TOTAL_FONT
ws6.cell(row=total_row_6a, column=3).alignment = center_align
ws6.cell(row=total_row_6a, column=5, value="=SUM(E5:E9)").font = TOTAL_FONT
ws6.cell(row=total_row_6a, column=5).alignment = center_align
for c in range(1, 7):
    ws6.cell(row=total_row_6a, column=c).fill = TOTAL_FILL
    ws6.cell(row=total_row_6a, column=c).border = thin_border
    ws6.cell(row=total_row_6a, column=c).alignment = center_align

# 更新 6C 盈亏汇总（行号变了，从原来的行+3开始）
# 原来 6C 在 summary_start 行，现在需要重新计算
# 6B 收入区域在 负债合计行后+2
income_start = total_row_6a + 2  # 行13
# 6C 在收入合计行后+2
summary_start = income_start + 5  # 行18

# 更新估算净盈亏
ws6.cell(row=summary_start + 4, column=1, value="预估净盈亏").font = TOTAL_FONT
ws6.cell(row=summary_start + 4, column=2, value="+1,270 ~ +3,588 元").font = TOTAL_FONT
ws6.cell(row=summary_start + 4, column=2).fill = GREEN_FILL

print("✅ Sheet 6 收支总账：已补全支出明细（光大+宁波）")

# ================================================
# 保存
# ================================================
wb.save(input_path)
print("✅ 信用卡主控表.xlsx 已保存")
print("   权益数据已同步：补全邮储高铁贵宾厅 + 更新预估收入为 3,150-5,468 元")
print("   收支总账已同步：支出明细补全光大500+宁波1800")