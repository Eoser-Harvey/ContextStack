"""年费到期提醒脚本 — 检查 Excel 中年费截止日期，支持邮件/桌面/微信/电话通知

=== 通知方式对比 ===
  ┌──────────┬────────────┬──────────────────────────────────┐
  │ 方式     │ 成本       │ 说明                             │
  ├──────────┼────────────┼──────────────────────────────────┤
  │ 桌面弹窗 │ 免费       │ 电脑上弹出通知（无需配置）        │
  │ 邮件通知 │ 免费       │ QQ邮箱即可，需配置SMTP授权码      │
  │ 微信通知 │ 免费       │ 通过 PushPlus/Server酱 推送       │
  │ 电话通知 │ 0.06元/通  │ 腾讯云语音通知，个人可开通        │
  │ 短信通知 │ 0.045元/条 │ 阿里云短信，需企业认证             │
  └──────────┴────────────┴──────────────────────────────────┘

=== 快速开始 ===
  1. 直接运行: python 年费到期提醒.py
     → 控制台输出 + 桌面弹窗（默认启用）
  2. 配置邮件: 修改下方 SMTP_CONFIG
  3. 配置微信: 修改下方 PUSHPLUS_TOKEN
  4. 每天自动运行: 运行 setup_task.ps1（管理员PowerShell）

=== Windows 定时任务（每天9:00自动运行）===
  管理员 PowerShell 执行:
  $action = New-ScheduledTaskAction -Execute "python" -Argument "年费到期提醒.py" -WorkingDirectory "当前目录"
  $trigger = New-ScheduledTaskTrigger -Daily -At "09:00"
  Register-ScheduledTask -TaskName "信用卡年费提醒" -Action $action -Trigger $trigger
"""

import openpyxl
import smtplib
import datetime
import re
import json
import subprocess
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path

# ============================================
# 配置区
# ============================================

# 提醒阈值：截止日期前多少天开始提醒
ALERT_DAYS = 30

# 邮件通知（QQ邮箱）
SMTP_CONFIG = {
    "enabled": False,                                    # 改为 True 启用
    "smtp_server": "smtp.qq.com",
    "smtp_port": 587,
    "sender_email": "你的QQ号@qq.com",
    "sender_password": "你的QQ邮箱授权码",                 # 设置→账户→POP3/SMTP→生成授权码
    "receiver_email": "你的QQ号@qq.com",
}

# 微信通知（PushPlus，免费，无需安装App）
# 注册获取 token: http://www.pushplus.plus/
PUSHPLUS_CONFIG = {
    "enabled": True,                                    # 改为 True 启用
    "token": "6b280503d6df417f87556a366bc4ba61",
}

# 电话通知（腾讯云语音通知，约0.06元/通，个人可开通）
# 开通步骤见下方注释
TENCENT_VOICE_CONFIG = {
    "enabled": False,                                   # 改为 True 启用
    "secret_id": "你的腾讯云SecretId",
    "secret_key": "你的腾讯云SecretKey",
    "app_id": "你的语音应用ID",                           # 腾讯云语音消息→应用管理→应用ID
    "template_id": "你的语音模板ID",                      # 语音模板，需审核通过
    "called_number": "你的手机号",                        # 接收电话的手机号（+86开头）
}

# 桌面弹窗（Windows 自带，无需配置）
DESKTOP_NOTIFY_ENABLED = True

# Excel 文件路径
EXCEL_PATH = Path(__file__).parent / "信用卡主控表.xlsx"

# ============================================
# 核心逻辑
# ============================================

def parse_deadline(text):
    """从截止日期文本中解析出 (年, 月, 日)"""
    if not text:
        return None
    text = str(text)
    m = re.search(r'(\d{1,2})月(\d{1,2})日', text)
    if m:
        month = int(m.group(1))
        day = int(m.group(2))
        return (datetime.date.today().year, month, day)
    m = re.search(r'(\d{1,2})月', text)
    if m:
        month = int(m.group(1))
        return (datetime.date.today().year, month, 1)
    return None

def check_annual_fee_deadlines(excel_path):
    """检查年费追踪Sheet中的到期日期"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["2-年费追踪"]
    today = datetime.date.today()
    alerts = []

    for row in range(1, ws.max_row + 1):
        bank = ws.cell(row=row, column=1).value
        card = ws.cell(row=row, column=2).value
        status = ws.cell(row=row, column=8).value
        
        if not bank or not card:
            continue
        if str(status) in ("✅", "—", "None"):
            continue
        
        for col in range(3, 8):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val and re.search(r'\d{1,2}月', str(cell_val)):
                deadline = parse_deadline(cell_val)
                if deadline:
                    deadline_date = datetime.date(deadline[0], deadline[1], deadline[2])
                    days_left = (deadline_date - today).days
                    if days_left < -90:
                        deadline_date = datetime.date(today.year + 1, deadline[1], deadline[2])
                        days_left = (deadline_date - today).days
                    
                    if 0 <= days_left <= ALERT_DAYS:
                        alerts.append({
                            "bank": bank, "card": card,
                            "deadline": str(cell_val),
                            "deadline_date": deadline_date.strftime("%Y-%m-%d"),
                            "days_left": days_left,
                            "status": str(status) if status else "未知",
                        })
                    elif days_left < 0 and days_left > -90:
                        alerts.append({
                            "bank": bank, "card": card,
                            "deadline": str(cell_val),
                            "deadline_date": deadline_date.strftime("%Y-%m-%d"),
                            "days_left": days_left,
                            "status": f"⚠️ 已过期{abs(days_left)}天",
                        })
    wb.close()
    return alerts

# ============================================
# 通知渠道
# ============================================

def send_desktop_notification(alerts):
    """Windows 桌面弹窗通知"""
    if not alerts:
        return
    try:
        # 使用 PowerShell 发送 Windows Toast 通知
        alert_text = "; ".join(
            f"{a['bank']}{a['card']} {a['days_left']}天后到期"
            for a in alerts[:3]  # 最多显示3条
        )
        if len(alerts) > 3:
            alert_text += f" 等{len(alerts)}条"
        
        ps_script = f'''
        [Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType = WindowsRuntime] > $null
        $template = [Windows.UI.Notifications.ToastNotificationManager]::GetTemplateContent([Windows.UI.Notifications.ToastTemplateType]::ToastText02)
        $textNodes = $template.GetElementsByTagName("text")
        $textNodes.Item(0).AppendChild($template.CreateTextNode("⚠️ 信用卡年费到期提醒")) > $null
        $textNodes.Item(1).AppendChild($template.CreateTextNode("{alert_text}")) > $null
        $toast = [Windows.UI.Notifications.ToastNotification]::new($template)
        [Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier("信用卡提醒").Show($toast)
        '''
        subprocess.run(["powershell", "-Command", ps_script], capture_output=True)
        print("✅ 桌面弹窗已发送")
    except Exception as e:
        print(f"⚠️ 桌面弹窗失败: {e}")

def send_email(alerts):
    """邮件通知"""
    if not SMTP_CONFIG["enabled"]:
        return
    subject = f"⚠️ 信用卡年费提醒 — {len(alerts)}张卡需要关注"
    
    lines = [
        f"<h2>信用卡年费到期提醒 — {datetime.date.today().strftime('%Y年%m月%d日')}</h2>",
        "<hr>",
        "<table border='1' cellpadding='8' cellspacing='0' style='border-collapse:collapse'>",
        "<tr style='background:#4472C4;color:white'><th>银行</th><th>卡种</th><th>截止日期</th><th>剩余天数</th><th>状态</th></tr>",
    ]
    for a in alerts:
        color = "#FF0000" if a["days_left"] < 0 else ("#FF8C00" if a["days_left"] <= 7 else "#000")
        bg = "#FFC7CE" if a["days_left"] < 0 else ("#FFCC99" if a["days_left"] <= 7 else "#FFF")
        lines.append(
            f"<tr style='background:{bg}'><td>{a['bank']}</td><td>{a['card']}</td>"
            f"<td>{a['deadline']}</td><td style='color:{color};font-weight:bold'>{a['days_left']}天</td>"
            f"<td>{a['status']}</td></tr>"
        )
    lines.append("</table><hr><p style='color:#808080;font-size:12px'>信用卡年费提醒系统自动发送</p>")
    
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = SMTP_CONFIG["sender_email"]
    msg["To"] = SMTP_CONFIG["receiver_email"]
    msg.attach(MIMEText("\n".join(lines), "html", "utf-8"))
    
    try:
        with smtplib.SMTP(SMTP_CONFIG["smtp_server"], SMTP_CONFIG["smtp_port"]) as server:
            server.starttls()
            server.login(SMTP_CONFIG["sender_email"], SMTP_CONFIG["sender_password"])
            server.sendmail(SMTP_CONFIG["sender_email"], SMTP_CONFIG["receiver_email"], msg.as_string())
        print(f"✅ 邮件已发送到 {SMTP_CONFIG['receiver_email']}")
    except Exception as e:
        print(f"❌ 邮件发送失败: {e}")

def send_pushplus(alerts):
    """微信通知（PushPlus）"""
    if not PUSHPLUS_CONFIG["enabled"]:
        return
    import urllib.request
    text = "\n".join(
        f"{a['bank']}{a['card']} | {a['deadline']} | 剩余{a['days_left']}天 | {a['status']}"
        for a in alerts
    )
    data = json.dumps({
        "token": PUSHPLUS_CONFIG["token"],
        "title": f"⚠️ 信用卡年费提醒 — {len(alerts)}条",
        "content": text,
        "template": "txt",
    }).encode()
    try:
        req = urllib.request.Request("http://www.pushplus.plus/send", data=data,
            headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req)
        print("✅ 微信通知已发送（PushPlus）")
    except Exception as e:
        print(f"❌ 微信通知失败: {e}")

def send_voice_call(alerts):
    """电话通知（腾讯云语音通知 VMS）"""
    if not TENCENT_VOICE_CONFIG["enabled"]:
        return
    try:
        import hashlib
        import hmac
        import time
        import urllib.request

        cfg = TENCENT_VOICE_CONFIG
        secret_id = cfg["secret_id"]
        secret_key = cfg["secret_key"]
        app_id = cfg["app_id"]
        template_id = cfg["template_id"]
        called_number = cfg["called_number"]

        # 语音模板参数：银行名+卡种+剩余天数（模板需提前在腾讯云审核）
        # 模板示例：您好，您的{1}年费即将到期，剩余{2}天，请及时处理。
        voice_params = []
        for a in alerts[:3]:  # 最多播报3条
            voice_params.append(f"{a['bank']}{a['card']} {a['days_left']}天")

        # 构建请求
        timestamp = int(time.time())
        params = {
            "Action": "SendTtsVoice",
            "Version": "2020-02-10",
            "Region": "ap-guangzhou",
            "Timestamp": timestamp,
            "Nonce": timestamp,
            "SecretId": secret_id,
            "TemplateId": template_id,
            "CalledNumber": called_number,
            "VoiceSdkAppid": app_id,
            "TemplateParamSet": voice_params,
            "PlayTimes": 2,
        }

        # 签名（腾讯云 API 3.0 签名）
        # 注意：实际使用需要完整的签名 V3 实现，此处为简化示例
        # 推荐安装 tencentcloud-sdk-python: pip install tencentcloud-sdk-python-vms
        print("💡 电话通知需要安装 tencentcloud-sdk-python-vms")
        print("   pip install tencentcloud-sdk-python-vms")
        print("   详细接入文档：https://cloud.tencent.com/document/product/1128")

        # 简化版：使用腾讯云 SDK
        try:
            from tencentcloud.common import credential
            from tencentcloud.common.profile.client_profile import ClientProfile
            from tencentcloud.common.profile.http_profile import HttpProfile
            from tencentcloud.vms.v20200902 import vms_client, models

            cred = credential.Credential(secret_id, secret_key)
            httpProfile = HttpProfile()
            httpProfile.endpoint = "vms.tencentcloudapi.com"
            clientProfile = ClientProfile()
            clientProfile.httpProfile = httpProfile
            client = vms_client.VmsClient(cred, "ap-guangzhou", clientProfile)

            req = models.SendTtsVoiceRequest()
            req.TemplateId = template_id
            req.CalledNumber = called_number
            req.VoiceSdkAppid = app_id
            req.TemplateParamSet = [f"{a['bank']}{a['card']} {a['days_left']}天到期" for a in alerts[:3]]
            req.PlayTimes = 2

            resp = client.SendTtsVoice(req)
            print(f"✅ 电话通知已发起（腾讯云语音）: {resp.SendStatus}")
        except ImportError:
            print("⚠️ 未安装 tencentcloud-sdk-python-vms，电话通知跳过")
            print("   安装命令: pip install tencentcloud-sdk-python-vms")
        except Exception as e:
            # 如果 SDK 不可用，使用原始 HTTP 请求
            print(f"⚠️ 电话通知 SDK 调用失败: {e}")
            print("   请确认已安装: pip install tencentcloud-sdk-python-vms")

    except Exception as e:
        print(f"❌ 电话通知失败: {e}")

# ============================================
# 主入口
# ============================================

def main():
    print(f"=== 信用卡年费到期提醒 === {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    if not EXCEL_PATH.exists():
        print(f"❌ Excel 文件不存在: {EXCEL_PATH}")
        return
    
    alerts = check_annual_fee_deadlines(EXCEL_PATH)
    
    if not alerts:
        print("✅ 暂无需要关注的年费到期提醒")
        return
    
    print(f"\n⚠️ 发现 {len(alerts)} 条需要关注的年费记录：\n")
    print("-" * 70)
    print(f"{'银行':<8} {'卡种':<12} {'截止日期':<12} {'剩余天数':<10} 状态")
    print("-" * 70)
    for a in alerts:
        flag = "🔴" if a["days_left"] < 0 else ("🟠" if a["days_left"] <= 7 else "🟡")
        print(f"{flag} {a['bank']:<6} {a['card']:<10} {a['deadline']:<12} {a['days_left']:<10}天 {a['status']}")
    print("-" * 70)
    
    # 桌面弹窗
    if DESKTOP_NOTIFY_ENABLED:
        send_desktop_notification(alerts)
    
    # 邮件
    if SMTP_CONFIG["enabled"]:
        send_email(alerts)
    else:
        print("\n💡 邮件通知未启用，修改 SMTP_CONFIG['enabled'] = True 并填入邮箱信息即可")
    
    # 微信
    if PUSHPLUS_CONFIG["enabled"]:
        send_pushplus(alerts)
    else:
        print("💡 微信通知未启用，注册 PushPlus 获取 token 即可推送微信消息")

    # 电话
    if TENCENT_VOICE_CONFIG["enabled"]:
        send_voice_call(alerts)
    else:
        print("💡 电话通知未启用，需开通腾讯云语音通知服务")

if __name__ == "__main__":
    main()